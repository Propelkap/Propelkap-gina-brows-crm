#!/usr/bin/env python3
"""
Migración de CITAS de Gina Brows desde reservas_GINABROWS.xlsx
Lee los 18 sheets, mapea contra clientes existentes y catálogo de servicios,
y hace UPSERT bulk a Supabase.

Uso:
    python3 scripts/migrate-citas.py [--dry-run]
"""
import os
import re
import sys
import json
from datetime import datetime
from collections import Counter
from urllib.request import Request, urlopen
from urllib.error import HTTPError

import openpyxl

DRY = "--dry-run" in sys.argv
SRC = "/Users/borrebriones/Downloads/reservas_GINABROWS.xlsx"

# Cargar envs desde .env.local
ENV_PATH = os.path.join(os.path.dirname(__file__), "..", ".env.local")
env = {}
with open(ENV_PATH) as f:
    for line in f:
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line: continue
        k, v = line.split("=", 1)
        env[k] = v.strip('"')

SUPABASE_URL = env["NEXT_PUBLIC_SUPABASE_URL"].rstrip("/")
SERVICE_KEY = env["SUPABASE_SERVICE_ROLE_KEY"]

HEADERS = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "resolution=merge-duplicates,return=representation",
}

# =========================================================================
# Mapeo de nombres de servicio (variantes en AgendaPro → catálogo)
# =========================================================================
SERVICIO_ALIAS = {
    "sesión microblading": "Microblading",
    "sesion microblading": "Microblading",
    "microblading": "Microblading",
    "retoque microblading": "Retoque mensual",
    "retoque mensual": "Retoque mensual",
    "retoque anual": "Retoque anual",
    "valoracion": "Valoración",
    "valoración": "Valoración",
    "diseño de ceja": "Diseño de ceja",
    "diseno de ceja": "Diseño de ceja",
    "1x sesión : remoción de tatuaje en cejas": "Remoción de ceja (1 sesión)",
    "1x sesion : remocion de tatuaje en cejas": "Remoción de ceja (1 sesión)",
    "3x sesiones : remoción de tatuaje en cejas": "Remoción de ceja (3 sesiones)",
    "3x sesiones : remocion de tatuaje en cejas": "Remoción de ceja (3 sesiones)",
    "1x sesion: hollywood peeling": "Hollywood peeling (1 sesión)",
    "1x sesión: hollywood peeling": "Hollywood peeling (1 sesión)",
    "3x sesiones: hollywood peeling": "Hollywood peeling (3 sesiones)",
    "3x sesiones hollywood peeling": "Hollywood peeling (3 sesiones)",
    "melasma": "Melasma",
    "henna brows": "Henna Brows",
    "depilación de ceja básica": "Depilación de Ceja Básica",
    "depilacion de ceja basica": "Depilación de Ceja Básica",
    "tocobo azul": "TOCOBO azul",
    "paquete mantenimiento de microblading": "Paquete Mantenimiento de Microblading",
    "láser express": "Láser express",
    "laser express": "Láser express",
}

# =========================================================================
# Mapeo de estados
# =========================================================================
ESTADO_MAP = {
    "asiste": "completada",
    "no asiste": "no_show",
    "cancelada": "cancelada",
    "cancela": "cancelada",
    "confirmado": "confirmada",
    "confirmada": "confirmada",
    "pendiente": "tentativa",
    "en espera": "tentativa",
    "reservado": "tentativa",
}

# =========================================================================
# Helpers
# =========================================================================
def clean(s):
    if s is None: return ""
    s = str(s)
    for c in ['\xa0', '‬', '​']:
        s = s.replace(c, '')
    return s.strip()

def normalize_phone(raw):
    p = re.sub(r'[^0-9+]', '', clean(raw))
    if not p: return None
    if not p.startswith("+"): p = "+" + p
    if len(p) < 10: return None
    return p

def normalize_email(raw):
    e = clean(raw).lower()
    if not e or "@" not in e: return None
    return e

def parse_dt(raw):
    """Parse '28/03/2024 10:00' or '28/03/2024' to ISO datetime."""
    s = clean(raw)
    if not s: return None
    # Try DD/MM/YYYY HH:MM
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})\s*(\d{1,2}):(\d{2})?', s)
    if m:
        d, mo, y, h, mi = m.groups()
        try:
            return datetime(int(y), int(mo), int(d), int(h), int(mi or 0)).isoformat() + "-06:00"
        except ValueError: return None
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if m:
        d, mo, y = m.groups()
        try:
            return datetime(int(y), int(mo), int(d), 11, 0).isoformat() + "-06:00"
        except ValueError: return None
    return None

def parse_price(raw):
    s = re.sub(r'[^\d.]', '', clean(raw))
    return float(s) if s else 0.0

def http_get(path):
    req = Request(f"{SUPABASE_URL}{path}", headers=HEADERS, method="GET")
    with urlopen(req) as r:
        return json.loads(r.read())

def http_post(path, body):
    data = json.dumps(body).encode()
    req = Request(f"{SUPABASE_URL}{path}", data=data, headers=HEADERS, method="POST")
    try:
        with urlopen(req) as r:
            return json.loads(r.read())
    except HTTPError as e:
        return {"error": e.read().decode()}

# =========================================================================
# Cargar lookup de clientes y servicios desde BD
# =========================================================================
print("📥 Cargando catálogo de servicios desde BD…")
servicios = http_get("/rest/v1/servicios?select=id,nombre,duracion_min,precio_mxn")
servicio_id_by_name = {s["nombre"]: s for s in servicios}
print(f"   {len(servicios)} servicios en catálogo")

print("📥 Cargando clientes desde BD para matching…")
clientes = http_get("/rest/v1/clientes?select=id,email,whatsapp_normalizado&limit=2000")
cliente_by_email = {c["email"]: c["id"] for c in clientes if c.get("email")}
cliente_by_phone = {c["whatsapp_normalizado"]: c["id"] for c in clientes if c.get("whatsapp_normalizado")}
print(f"   {len(clientes)} clientes ({len(cliente_by_email)} con email, {len(cliente_by_phone)} con tel)")

# =========================================================================
# Leer todos los sheets
# =========================================================================
print(f"\n📥 Leyendo {SRC}…")
wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

todas_citas = []
sin_servicio = Counter()
sin_cliente = []
nuevos_clientes_de_citas = {}  # email/phone → datos para crear si no existe

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows: continue
    h = rows[0]
    if "Servicio" not in h: continue

    idx = {col: h.index(col) for col in [
        "Fecha de realización", "Fecha de creación", "Nombre", "Apellido", "E-mail", "Teléfono",
        "Servicio", "Precio lista", "Precio real", "Estado", "ID pago",
        "Comentario interno", "Notas compartidas con cliente"
    ] if col in h}

    for r in rows[1:]:
        if not any(r): continue
        servicio_raw = clean(r[idx["Servicio"]])
        if not servicio_raw: continue
        servicio_key = servicio_raw.lower()
        servicio_nombre = SERVICIO_ALIAS.get(servicio_key)
        if not servicio_nombre or servicio_nombre not in servicio_id_by_name:
            sin_servicio[servicio_raw] += 1
            continue
        servicio = servicio_id_by_name[servicio_nombre]

        fecha_iso = parse_dt(r[idx["Fecha de realización"]])
        if not fecha_iso: continue

        # Match cliente
        email = normalize_email(r[idx["E-mail"]])
        phone = normalize_phone(r[idx["Teléfono"]])
        cliente_id = (email and cliente_by_email.get(email)) or (phone and cliente_by_phone.get(phone))

        if not cliente_id:
            # Cliente no existe — guardar para crear (ej. clientas que solo aparecen en citas)
            nombre_cita = clean(r[idx.get("Nombre", -1)] if "Nombre" in idx else "")
            apellido_cita = clean(r[idx.get("Apellido", -1)] if "Apellido" in idx else "")
            key = email or phone
            if key and key not in nuevos_clientes_de_citas:
                nuevos_clientes_de_citas[key] = {
                    "nombre": nombre_cita or "(sin nombre)",
                    "apellido": apellido_cita or None,
                    "email": email,
                    "whatsapp": phone,
                    "agendapro_id": f"ap_{key}",
                    "estado": "activa",
                    "migrado_desde_agendapro_at": datetime.now().isoformat(),
                }
            sin_cliente.append({"nombre": f"{nombre_cita} {apellido_cita}", "email": email, "phone": phone})
            continue

        precio = parse_price(r[idx.get("Precio real", -1)] if "Precio real" in idx else 0)
        if precio == 0:
            precio = parse_price(r[idx.get("Precio lista", -1)] if "Precio lista" in idx else 0) or servicio["precio_mxn"]

        estado_raw = clean(r[idx["Estado"]]).lower()
        estado = ESTADO_MAP.get(estado_raw, "completada")

        # Calcular fin con duración del servicio
        from datetime import timedelta
        inicio_dt = datetime.fromisoformat(fecha_iso[:-6])  # quitar tz para sumar
        fin_dt = inicio_dt + timedelta(minutes=servicio["duracion_min"])
        fin_iso = fin_dt.isoformat() + "-06:00"

        notas = clean(r[idx.get("Comentario interno", -1)] if "Comentario interno" in idx else "") or None

        # ID único de cita — preferir ID pago de AgendaPro si existe (único garantizado),
        # si no, combo único con fecha de creación de la cita en AgendaPro
        id_pago = clean(r[idx["ID pago"]]) if "ID pago" in idx else ""
        fecha_creacion = parse_dt(r[idx["Fecha de creación"]]) if "Fecha de creación" in idx else None
        if id_pago and id_pago != "NA":
            ag_id = f"ap_pago_{id_pago}"
        elif fecha_creacion:
            ag_id = f"ap_{cliente_id}_{fecha_iso[:16]}_{servicio['id'][:8]}_{fecha_creacion[:19]}"
        else:
            # último fallback: usar todo el datetime + servicio + ya tendrá un sufijo abajo si se duplica
            ag_id = f"ap_{cliente_id}_{fecha_iso[:19]}_{servicio['id'][:8]}"

        todas_citas.append({
            "cliente_id": cliente_id,
            "servicio_id": servicio["id"],
            "inicio": fecha_iso,
            "fin": fin_iso,
            "estado": estado,
            "precio_mxn": float(precio),
            "notas_internas": notas,
            "agendapro_id": ag_id,
        })

# =========================================================================
# Reporte pre-migración
# =========================================================================
print(f"\n📊 ANÁLISIS DE CITAS")
print(f"   Total citas válidas:        {len(todas_citas)}")
print(f"   Sin cliente match:          {len(sin_cliente)}")
print(f"   Clientas nuevas a crear:    {len(nuevos_clientes_de_citas)}")
if sin_servicio:
    print(f"\n⚠️  Servicios SIN MAPEO ({sum(sin_servicio.values())} citas afectadas):")
    for s, n in sin_servicio.most_common():
        print(f"   {n:5d}  {s!r}")

# =========================================================================
# Migrar
# =========================================================================
if DRY:
    print("\n🟡 DRY-RUN: nada se escribió en BD.")
    print("\nPrimeras 3 citas que serían insertadas:")
    for c in todas_citas[:3]:
        print(f"   {c['inicio']} | servicio_id={c['servicio_id'][:8]}… | ${c['precio_mxn']:.0f} | {c['estado']}")
    sys.exit(0)

# 1. Insertar clientes nuevos (las que aparecen en citas pero no en clientes export)
if nuevos_clientes_de_citas:
    print(f"\n🆕 Insertando {len(nuevos_clientes_de_citas)} clientes nuevos detectados en citas…")
    nuevos = list(nuevos_clientes_de_citas.values())
    BATCH = 100
    for i in range(0, len(nuevos), BATCH):
        result = http_post("/rest/v1/clientes?on_conflict=agendapro_id", nuevos[i:i+BATCH])
        if isinstance(result, dict) and "error" in result:
            print(f"   ❌ Error batch {i}: {result['error'][:200]}")
        else:
            print(f"   ✓ Batch {i}: {len(result)} insertados")
    # Recargar lookup
    clientes = http_get("/rest/v1/clientes?select=id,email,whatsapp_normalizado&limit=2000")
    cliente_by_email = {c["email"]: c["id"] for c in clientes if c.get("email")}
    cliente_by_phone = {c["whatsapp_normalizado"]: c["id"] for c in clientes if c.get("whatsapp_normalizado")}

# Re-procesar citas sin_cliente con el lookup actualizado
if sin_cliente and nuevos_clientes_de_citas:
    print(f"\n🔁 Re-mapeando {len(sin_cliente)} citas que no tenían cliente…")
    nuevas_citas = []
    for c in sin_cliente:
        cid = (c["email"] and cliente_by_email.get(c["email"])) or (c["phone"] and cliente_by_phone.get(c["phone"]))
        if cid:
            # No tenemos los datos completos de la cita acá; este path se simplifica
            # porque ya guardamos las citas en `todas_citas` solo si tenían cliente.
            # Para que el re-match funcione completo, re-procesamos el archivo entero.
            pass
    # Truco: re-procesar el archivo entero ahora que los nuevos clientes están
    print("   (Re-procesando archivo completo para enganchar clientes nuevos…)")
    # Reset y volver a leer
    todas_citas_ronda2 = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows: continue
        h = rows[0]
        if "Servicio" not in h: continue
        idx = {col: h.index(col) for col in [
            "Fecha de realización", "Fecha de creación", "E-mail", "Teléfono", "Servicio",
            "Precio lista", "Precio real", "Estado", "ID pago", "Comentario interno"
        ] if col in h}
        for r in rows[1:]:
            if not any(r): continue
            servicio_raw = clean(r[idx["Servicio"]])
            servicio_nombre = SERVICIO_ALIAS.get(servicio_raw.lower())
            if not servicio_nombre or servicio_nombre not in servicio_id_by_name: continue
            servicio = servicio_id_by_name[servicio_nombre]
            fecha_iso = parse_dt(r[idx["Fecha de realización"]])
            if not fecha_iso: continue
            email = normalize_email(r[idx["E-mail"]])
            phone = normalize_phone(r[idx["Teléfono"]])
            cliente_id = (email and cliente_by_email.get(email)) or (phone and cliente_by_phone.get(phone))
            if not cliente_id: continue
            precio = parse_price(r[idx.get("Precio real", -1)]) or parse_price(r[idx.get("Precio lista", -1)]) or servicio["precio_mxn"]
            estado_raw = clean(r[idx["Estado"]]).lower()
            estado = ESTADO_MAP.get(estado_raw, "completada")
            from datetime import timedelta
            inicio_dt = datetime.fromisoformat(fecha_iso[:-6])
            fin_iso = (inicio_dt + timedelta(minutes=servicio["duracion_min"])).isoformat() + "-06:00"
            notas = clean(r[idx.get("Comentario interno", -1)]) or None
            id_pago = clean(r[idx["ID pago"]]) if "ID pago" in idx else ""
            fecha_creacion = parse_dt(r[idx["Fecha de creación"]]) if "Fecha de creación" in idx else None
            if id_pago and id_pago != "NA":
                ag_id = f"ap_pago_{id_pago}"
            elif fecha_creacion:
                ag_id = f"ap_{cliente_id}_{fecha_iso[:16]}_{servicio['id'][:8]}_{fecha_creacion[:19]}"
            else:
                ag_id = f"ap_{cliente_id}_{fecha_iso[:19]}_{servicio['id'][:8]}"
            todas_citas_ronda2.append({
                "cliente_id": cliente_id,
                "servicio_id": servicio["id"],
                "inicio": fecha_iso,
                "fin": fin_iso,
                "estado": estado,
                "precio_mxn": float(precio),
                "notas_internas": notas,
                "agendapro_id": ag_id,
            })
    todas_citas = todas_citas_ronda2
    print(f"   Citas re-procesadas: {len(todas_citas)}")

# 2. Dedupe defensivo: si quedan colisiones, agregar sufijo numérico
seen_ids = {}
deduped = []
colisiones = 0
for c in todas_citas:
    base = c["agendapro_id"]
    if base in seen_ids:
        seen_ids[base] += 1
        c["agendapro_id"] = f"{base}_dup{seen_ids[base]}"
        colisiones += 1
    else:
        seen_ids[base] = 0
    deduped.append(c)
todas_citas = deduped
if colisiones:
    print(f"\n⚠️  Resueltas {colisiones} colisiones de agendapro_id con sufijo")

# Insertar citas en batch
print(f"\n🔄 Insertando {len(todas_citas)} citas en BD…")
BATCH = 200
inserted = 0
for i in range(0, len(todas_citas), BATCH):
    batch = todas_citas[i:i+BATCH]
    result = http_post("/rest/v1/citas?on_conflict=agendapro_id", batch)
    if isinstance(result, dict) and "error" in result:
        print(f"   ❌ Error batch {i}: {result['error'][:300]}")
    else:
        inserted += len(result) if isinstance(result, list) else 0
        sys.stdout.write(f"\r   {i + len(batch)}/{len(todas_citas)}")
        sys.stdout.flush()
print(f"\n   ✅ Insertadas/actualizadas: {inserted}")

# 3. Marcar dormidas
print(f"\n🔧 Calculando dormidas…")
result = http_post("/rest/v1/rpc/marcar_dormidas", {})
print(f"   ✅ Dormidas marcadas: {result}")

print(f"\n✨ Listo.\n")
